import json
import os
from datetime import datetime
from random import random

from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect
from django.views.decorators.csrf import csrf_exempt

from app01 import models
from app01.utils.pagination import Pagination
from django import forms
from app01.utils.bootstrap import BootStrapModelForm
from app01.utils.bootstrap import BootStrapForm
from app01.utils.encrypt import md5
from app01.utils.code import check_code
from django.conf import settings


# from django.utils.safestring import mark_safe
# Create your views here.


def depart_list(request):
    # 部门列表
    # 去数据库中获取列表
    # 这个获取到的是类似于二维数组那样的形式，一行一行，每一行包括每个字段的数据
    queryset = models.Department.objects.all()
    page_object = Pagination(request, queryset)
    context = {
        "queryset": page_object.page_queryset,  # 分完页的数据
        "page_string": page_object.html()  # 生成页码
    }
    return render(request, 'depart_list.html', context)


def depart_add(request):
    if request.method == "GET":
        return render(request, 'depart_add.html')

    # 获取用户通过post提交的数据 (title输入为空)
    title = request.POST.get("title")

    # 保存到数据库
    models.Department.objects.create(title=title)

    # 重定向回部门列表
    return redirect("/depart/list/")


def depart_delete(request):
    # 删除部门
    # 通过url连接传参的形式读取数据 127.0.0.1/……/?nid = 1 像这样
    nid = request.GET.get('nid')
    # 删除
    models.Department.objects.filter(id=nid).delete()

    # 重新定向回列表页面
    return redirect("/depart/list/")


#                       然后这边需要接受那个url里传来的数据
def depart_edit(request, nid):
    # 修改部门
    if request.method == "GET":
        # 根据获取的nid，获取数据，但这样获取的是一个里面只有一个对象的数组
        row_object = models.Department.objects.filter(id=nid).first()
        return render(request, "depart_edit.html", {"默认值": row_object.title})

    title = request.POST.get("title")  # 获取数据

    # 根据id进行更新
    models.Department.objects.filter(id=nid).update(title=title)
    return redirect("/depart/list/")


def user_list(request):
    # 用户管理
    queryset = models.UserInfo.objects.all()
    # 2.实例化分页对象
    page_object = Pagination(request, queryset)
    context = {
        "queryset": page_object.page_queryset,  # 分完页的数据
        "page_string": page_object.html()  # 生成页码
    }
    return render(request, 'user_list.html', context)


############### modle form 实例
class UserModelForm(BootStrapModelForm):
    name = forms.CharField(min_length=3, label="用户名")  # 字段想要限制条件就需要单独写，写法同models

    # password = forms.CharField(label="密码")
    class Meta:
        model = models.UserInfo  # 注意这里不用写depart_id，只要写depart就行了
        fields = ["name", "password", "age", "account", "create_time", "gender", "depart"]
    #     # widgets = {  # 插件，可以设置细节样式
    #     #     "name": forms.TextInput(attrs={"class": "form-control"}),   # 输入框，attrs里面写细节
    #     #     "password": forms.PasswordInput(attrs={"class": "form-control"}), # 输入框，attrs里面写细节
    #     #     "age": forms.TextInput(attrs={"class": "form-control"}), # 输入框，attrs里面写细节
    #     #     "account": forms.TextInput(attrs={"class": "form-control"}), # 输入框，attrs里面写细节
    #     #     "create_time": forms.TextInput(attrs={"id": "datepicker", "class": "form-control", "data-provide": "datepicker"}),
    #     #     # 输入框，attrs里面写细节
    #     #     "gender": forms.TextInput(attrs={"class": "form-control"}), # 输入框，attrs里面写细节
    #     #     "depart": forms.TextInput(attrs={"class": "form-control"}), # 输入框，attrs里面写细节
    #     # }
    # # 这边的缩进务必要看清
    # def __init__(self, *args, **kwargs):
    #     super().__init__(*args, **kwargs)
    #     for name, field in self.fields.items():
    #         # print(name, field)
    #         field.widget.attrs = {"class": "form-control", "placeholder": field.label}


def user_model_form_add(request):
    # 添加用户，model form版本
    if request.method == "GET":
        form = UserModelForm()
        return render(request, 'user_model_form_add.html', {"form": form})

    # 用户通过POST提交数据
    form = UserModelForm(data=request.POST)  # 拿到提交的所有数据
    if form.is_valid():
        # 如果合法，保存到数据库
        form.save()  # 定义的是那个类就自动存储到哪个类里面
        return redirect('/user/list/')
        # print(form.cleaned_data)

    return render(request, 'user_model_form_add.html', {"form": form})


def user_edit(request, nid):
    #     编辑用户
    #   根据id去获取那个对象
    row_object = models.UserInfo.objects.filter(id=nid).first()
    if request.method == "GET":
        form = UserModelForm(instance=row_object)
        return render(request, 'user_edit.html', {"form": form})

    #                                       这样就可以使获取的数据是指定那一行的
    form = UserModelForm(data=request.POST, instance=row_object)  # 获取post传来的数据
    if form.is_valid():
        form.save()  # 所以保存时就是修改哪一行的数据，而不是新增了
        return redirect('/user/list/')

    return render(request, 'user_edit.html', {"form": form})


def user_delate(request, nid):
    models.UserInfo.objects.filter(id=nid).delete()
    return redirect('/user/list/')


def pretty_list(request):
    # 靓号列表

    # page_size = 10
    # page = int(request.GET.get('page', 1)) # 如果没有就是第一页
    # start = (page - 1) * page_size
    # end = page * page_size
    # # for i in range(300):
    # #     models.PrettyNum.objects.create(mobile="18188888881", price=10, level=1, stauts=1)
    data_dict = {}
    search_data = request.GET.get('q', "")  # 说明它只要没有q就是一个空字符串
    if search_data:
        data_dict["mobile__contains"] = search_data
    # data_count = models.PrettyNum.objects.filter(**data_dict).count()
    # 包含关键词    降序排序
    queryset = models.PrettyNum.objects.filter(**data_dict).order_by("-level")
    # page_str_list = []
    # # 两个返回值，第一个商，第二个余数
    # a, b=divmod(data_count, page_size)
    # if b:
    #     a += 1
    # # for循环range是 [a, b)所以要 +1
    # plus = 5
    # if a <= 2 * plus + 1:
    #     start_page = 1
    #     end_page = a + 1
    # else:
    #     if page <= plus:
    #         start_page = 1
    #         end_page = 2 * plus + 2
    #     else:
    #         if (page + plus) > a:
    #             start_page = a - 2 * plus
    #             end_page = a + 1
    #         else:
    #             start_page = page - plus
    #             end_page = page + plus + 1
    #
    # prev = '<li><a href="?page={}">首页</a></li>'.format(1)
    # page_str_list.append(prev)
    #
    # if page > 1:
    #     prev = '<li><a href="?page={}">上一页</a></li>'.format(page - 1)
    # else:
    #     prev = '<li><a href="?page={}">上一页</a></li>'.format(1)
    #
    # page_str_list.append(prev)
    # for i in range(start_page, end_page):
    #     if i == page:
    #         ele = '<li class="active"><a href="?page={}">{}</a></li>'.format(i, i)
    #     else:
    #         ele = '<li><a href="?page={}">{}</a></li>'.format(i, i)
    #
    #     page_str_list.append(ele)
    #
    #
    # if page < a:
    #     nextt = '<li><a href="?page={}">下一页</a></li>'.format(page + 1)
    # else:
    #     nextt = '<li><a href="?page={}">下一页</a></li>'.format(a)
    # page_str_list.append(nextt)
    # nextt = '<li><a href="?page={}">尾页</a></li>'.format(a)
    # page_str_list.append(nextt)
    # page_string = mark_safe("".join(page_str_list))
    page_object = Pagination(request, queryset)

    context = {
        "queryset": page_object.page_queryset,  # 分完页的数据
        "page_string": page_object.html()  # 生成页码
    }

    # return render(request, "pretty_list.html",{"queryset": queryset, "search_data": search_data, "page_string": page_string})
    return render(request, "pretty_list.html", context)


# 正则表达式的引入
# from django.core.validators import RegexValidator
from django.core.exceptions import ValidationError


class PrettyModelForm(BootStrapModelForm):
    # 验证方法一：
    # mobile = forms.CharField(
    #     label="手机号",
    #     #  正则表达式固定格式 条件（这是标准手机号的正则表达式）  不符合输出的
    #     validators=[RegexValidator(r'^1[3-9]\d{9}$', "手机号格式错误")],
    # )
    class Meta:
        model = models.PrettyNum
        fields = ["mobile", "price", "level", "stauts"]

    #     验证方法2：
    def clean_mobile(self):
        # 获取用户提交的数据
        txt_mobile = self.cleaned_data["mobile"]
        exists = models.PrettyNum.objects.filter(mobile=txt_mobile).exists()
        if exists:
            raise ValidationError("手机号已存在")
        if len(txt_mobile) != 11:
            #             验证不通过
            raise ValidationError("格式错误")
        #         验证通过，把用户输入的值返回
        return txt_mobile


def pretty_add(request):
    #         添加靓号
    if request.method == "GET":
        form = PrettyModelForm()
        return render(request, 'pretty_add.html', {"form": form})
    form = PrettyModelForm(data=request.POST)

    if form.is_valid():
        form.save()
        return redirect('/pretty/list/')
    return render(request, 'pretty_add.html', {"form": form})


class PrettyEditModelForm(BootStrapModelForm):
    # 定义字段不可修改，                        label是备注名
    # mobile = forms.CharField(label="手机号")
    class Meta:
        model = models.PrettyNum
        fields = ["mobile", "price", "level", "stauts"]

    #     验证方法2：
    def clean_mobile(self):
        # 获取用户提交的数据
        txt_mobile = self.cleaned_data["mobile"]
        exists = models.PrettyNum.objects.filter(mobile=txt_mobile).exclude(id=self.instance.pk).exists()
        if exists:
            raise ValidationError("手机号已存在")
        if len(txt_mobile) != 11:
            #             验证不通过
            raise ValidationError("格式错误")
        #         验证通过，把用户输入的值返回
        return txt_mobile


def pretty_edit(request, nid):
    # 编辑良好
    a = models.PrettyNum.objects.filter(id=nid).first()
    if request.method == "GET":
        form = PrettyEditModelForm(instance=a)
        #            加一个instance可以实现只要一行数据
        return render(request, 'pretty_edit.html', {"form": form})
    form = PrettyEditModelForm(data=request.POST, instance=a)
    if form.is_valid():
        form.save()
        return redirect('/pretty/list/')
    return render(request, 'pretty_edit.html', {"form": form})


def pretty_delate(request, nid):
    models.PrettyNum.objects.filter(id=nid).delete()
    return redirect('/pretty/list/')


def admin_list(request):
    """管理员列表"""
    # 登录校验

    # 构造搜索
    data_dict = {}
    search_data = request.GET.get('q', "")  # 说明它只要没有q就是一个空字符串
    if search_data:
        data_dict["username__contains"] = search_data
    queryset = models.Admin.objects.filter(**data_dict)

    page_object = Pagination(request, queryset)

    context = {
        "queryset": page_object.page_queryset,  # 分完页的数据
        "page_string": page_object.html(),  # 生成页码
        "search_data": search_data  # 这个传回去主要是为了在搜索框中有默认值
    }
    return render(request, 'admin_list.html', context)


class AdminModelForm(BootStrapModelForm):
    confirm_password = forms.CharField(
        label="确认密码",
        widget=forms.PasswordInput(render_value=True)
    )

    class Meta:
        model = models.Admin
        fields = ["username", "password", "confirm_password"]
        widgets = {  # 这样做是额外添加一些样式
            "password": forms.PasswordInput(render_value=True),
        }

    def clean_password(self):
        pwd = self.cleaned_data.get("password")
        return md5(pwd)

    #     校验两次密码一致
    def clean_confirm_password(self):
        pwd = self.cleaned_data.get("password")
        confirm = md5(self.cleaned_data.get("confirm_password"))
        if confirm != pwd:
            raise ValidationError("密码不一致")
        return confirm


def admin_add(request):
    # 添加管理员

    if request.method == "GET":
        form = AdminModelForm()  # 实例化
        return render(request, 'change.html', {"title": "添加管理员", "form": form})

    form = AdminModelForm(data=request.POST)
    if form.is_valid():
        form.save()
        return redirect('/admin/list/')

    return render(request, 'change.html', {"title": "添加管理员", "form": form})


class AdminEditModelForm(BootStrapModelForm):
    class Meta:
        model = models.Admin
        fields = ["username"]


def admin_edit(request, nid):
    #     编辑管理员
    title = "编辑管理员"
    # 对象/NONE
    row_object = models.Admin.objects.filter(id=nid).first()
    if not row_object:
        return render(request, 'error.html', {"msg": "用户不存在"})

    if request.method == "GET":
        form = AdminEditModelForm(instance=row_object)
        return render(request, 'change.html', {"title": "添加管理员", "form": form})

    #     post修改
    form = AdminEditModelForm(data=request.POST, instance=row_object)
    if form.is_valid():
        form.save()
        return redirect('/admin/list/')

    return render(request, 'change.html', {"title": "添加管理员", "form": form})


def admin_delete(request, nid):
    # 删除管理员
    models.Admin.objects.filter(id=nid).delete()
    return redirect('/admin/list/')


def aa(request):
    # 默认页面
    return redirect('/admin/list/')


class AdminResetModelForm(BootStrapModelForm):
    confirm_password = forms.CharField(  # 新建一个输入框的意思
        label="确认密码",
        widget=forms.PasswordInput(render_value=True)
    )

    class Meta:
        model = models.Admin
        fields = ['password']
        widgets = {
            "password": forms.PasswordInput(render_value=True)
        }

    def clean_password(self):
        pwd = self.cleaned_data.get("password")
        md5_pwd = md5(pwd)

        # 去数据库校验当前密码和之前的密码是否一致
        #                                函数传过来的，pk就是id
        exists = models.Admin.objects.filter(id=self.instance.pk, password=md5_pwd).exists()
        if exists:  # 如果已经存在
            raise ValidationError("密码不能和以前的相同")

        return md5(pwd)
        #     校验两次密码一致

    def clean_confirm_password(self):
        pwd = self.cleaned_data.get("password")
        confirm = md5(self.cleaned_data.get("confirm_password"))
        if confirm != pwd:
            raise ValidationError("密码不一致")
        return confirm


def admin_reset(request, nid):
    #     编辑管理员
    # 对象/NONE
    row_object = models.Admin.objects.filter(id=nid).first()
    title = "重置密码 - {}".format(row_object.username)
    if not row_object:
        return render(request, 'error.html', {"msg": "用户不存在"})

    if request.method == "GET":
        form = AdminResetModelForm()
        return render(request, 'change.html', {"title": title, "form": form})

    #     post修改，保存之前的这个接受数据的form要加instence
    form = AdminResetModelForm(data=request.POST, instance=row_object)
    if form.is_valid():
        form.save()
        return redirect('/admin/list/')

    return render(request, 'change.html', {"title": title, "form": form})


# form
class LoginForm(BootStrapForm):
    username = forms.CharField(
        label="用户名",
        widget=forms.TextInput(),
        required=True,  # 必填
    )
    password = forms.CharField(
        label="密码",
        widget=forms.PasswordInput(render_value=True),
        required=True,  # 必填
    )
    code = forms.CharField(
        label="验证码",
        widget=forms.TextInput(),
        required=True,  # 必填
    )

    def clean_password(self):
        pwd = self.cleaned_data.get("password")
        return md5(pwd)


def login(request):
    # 登录
    if request.method == "GET":
        form = LoginForm()
        return render(request, 'login.html', {"form": form})

    form = LoginForm(data=request.POST)
    if form.is_valid():
        # 验证码校验
        # 在验证成功信息中弹出code，方便下面查询
        user_input_code = form.cleaned_data.pop("code")
        # 获取session中正确的验证码
        code = request.session.get("image_code")
        if code.upper() != user_input_code.upper():
            form.add_error("code", "验证码错误")
            return render(request, 'login.html', {"form": form})

        admin_object = models.Admin.objects.filter(**form.cleaned_data).first()
        if not admin_object:
            form.add_error("password", "用户名密码错误")
            return render(request, 'login.html', {"form": form})
        # cookie
        request.session["info"] = {"id": admin_object.id, "name": admin_object.username}
        # session可以保存7天
        request.session.set_expiry(60 * 60 * 24 * 7)
        return redirect('/admin/list/')
    return render(request, 'login.html', {"form": form})


def logout(request):
    #     注销
    request.session.clear()  # 清除cookie
    return redirect('/login/')


from io import BytesIO


def image_code(request):
    # 生成图片验证码
    img, code_string = check_code()

    # 写入到session中以便于后续获取
    request.session["image_code"] = code_string
    # 在session中设置验证码60秒超时
    request.session.set_expiry(60)

    stream = BytesIO()
    img.save(stream, 'png')

    return HttpResponse(stream.getvalue())


from django.views.decorators.csrf import csrf_exempt


class TaskModelForm(BootStrapModelForm):
    class Meta:
        model = models.Task
        fields = "__all__"


@csrf_exempt
def task_list(request):
    # 任务列表
    queryset = models.Task.objects.all().order_by('-id')
    form = TaskModelForm()
    page_object = Pagination(request, queryset)
    context = {
        "form": form,
        "queryset": page_object.page_queryset,  # 分完页的数据
        "page_string": page_object.html()  # 生成页码
    }
    return render(request, "task_list.html", context)


@csrf_exempt
def task_add(request):
    #     添加任务
    print(request.POST)
    form = TaskModelForm(data=request.POST)
    if form.is_valid():
        form.save()
        data_dict = {"status": True}
        return HttpResponse(json.dumps(data_dict))
    data_dict = {"status": False, "error": form.errors}
    return HttpResponse(json.dumps(data_dict, ensure_ascii=False))


@csrf_exempt
def task_ajax(request):
    # ajax学习
    print(request.GET)  # 这样可以获取get数据
    print(request.POST)  # 这样可以获取post数据
    data_dict = {"status": True, 'data': [11, 22, 33, 44]}
    return HttpResponse(json.dumps(data_dict))


class OrderModelForm(BootStrapModelForm):
    class Meta:
        model = models.Order
        # fields = "__all__"
        # fields = [""]
        exclude = ["oid", "admin"]


def order_list(request):
    queryset = models.Order.objects.all().order_by('-id')
    page_object = Pagination(request, queryset)
    form = OrderModelForm()

    context = {
        'form': form,
        "queryset": page_object.page_queryset,  # 分完页的数据
        "page_string": page_object.html()  # 生成页码
    }

    return render(request, 'order_list.html', context)


import random
from datetime import datetime


@csrf_exempt
def order_add(request):
    # 新建订单
    form = OrderModelForm(data=request.POST)
    if form.is_valid():
        form.instance.admin_id = request.session["info"]["id"]
        form.instance.oid = datetime.now().strftime("%Y%m%d%H%M%S") + str(random.randint(1000, 9999))
        form.save()
        return JsonResponse({"status": True})
    return JsonResponse({"status": False, "error": form.errors})


def order_delete(request):
    uid = request.GET.get("uid")
    exists = models.Order.objects.filter(id=uid).exists()
    if not exists:
        return JsonResponse({"status": False, "error": "删除失败 数据不存在"})
    models.Order.objects.filter(id=uid).delete()
    return JsonResponse({"status": True})


def order_detail(request):
    # 根据id获取订单信息
    # 方式1：
    """
    uid = request.GET.get("uid")
    row_object = models.Order.objects.filter(id=uid).filter()
    if not row_object:
        return JsonResponse({"status": False, "error": "删除失败 数据不存在"})
    result = {
        "status": True,
        "data": {
            "title": row_object.title,
            "price": row_object.price,
            "status": row_object.status,
        }
    }
    return JsonResponse(result)
    """
    #     方法2：
    uid = request.GET.get("uid")
    row_dict = models.Order.objects.filter(id=uid).values("title", 'price', 'status').first()
    if not row_dict:
        return JsonResponse({"status": False, 'error': "数据不存在。"})

    # 从数据库中获取到一个对象 row_object
    result = {
        "status": True,
        "data": row_dict
    }
    return JsonResponse(result)


@csrf_exempt
def order_edit(request):
    # 编辑页面
    nid = request.GET.get("nid")
    row_object = models.Order.objects.filter(id=nid).first()
    if not row_object:
        # print(1)
        return JsonResponse({"status": False, "tips": "数据不存在, 请刷新重试"})

    form = OrderModelForm(data=request.POST, instance=row_object)
    if form.is_valid():
        form.save()
        return JsonResponse({"status": True})
    return JsonResponse({"status": False, "errors": form.errors})


def chart_list(request):
    # 统计页面
    return render(request, "chart_list.html")


def chart_bar(request):
    # 构造柱状图数据
    # 数据可以去数据库获取
    legend = ['销量', '业绩']
    series_list = [
        {
            "name": '销量',
            "type": 'bar',
            "data": [5, 20, 36, 10, 10, 20, 65]
        },
        {
            "name": '业绩',
            "type": 'bar',
            "data": [5, 25, 20, 15, 5, 25, 11]
        }
    ]
    x_axis = ['1月', '2月', '3月', '4月', '5月', '6月', '7月']
    result = {
        "status": True,
        "data": {
            "legend": legend,
            "series_list": series_list,
            "x_axis": x_axis,
        }
    }
    return JsonResponse(result)


def upload_list(request):
    if request.method == "GET":
        return render(request, 'upload_list.html')

    # # 'username': ['big666']
    # print(request.POST)  # 请求体中数据
    # # {'avatar': [<InMemoryUploadedFile: 图片 1.png (image/png)>]}>
    # print(request.FILES)  # 请求发过来的文件 {}

    file_object = request.FILES.get("avatar")  # 这样获取到传过来的文件
    # print(file_object.name)  # 文件名：WX20211117-222041@2x.png

    f = open(file_object.name, mode='wb')  # 然后打开一个文件
    for chunk in file_object.chunks():  # 循环遍历获取的那个文件
        f.write(chunk)  # 将文件一点一点的写入
    f.close()  # 关闭文件
    return HttpResponse("...")


def depart_multi(request):
    """ 批量删除（Excel文件）"""
    from openpyxl import load_workbook

    # 1.获取用户上传的文件对象
    file_object = request.FILES.get("exc")

    # 2.对象传递给openpyxl，由openpyxl读取文件的内容
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]  # 这是获取第一张表

    # 3.循环获取每一行数据
    for row in sheet.iter_rows(min_row=1):  # 行数限制，从第二行开始
        text = row[0].value  # 第一列的值
        exists = models.Department.objects.filter(title=text).exists()  # 判断是否存在
        if not exists:
            models.Department.objects.create(title=text)  # 保存第一列的值

    return redirect('/depart/list/')


class UpForm(BootStrapForm):
    name = forms.CharField(label="姓名")
    age = forms.IntegerField(label="年龄")
    img = forms.FileField(label="头像")


def upload_form(request):
    title = "Form上传"
    if request.method == "GET":
        form = UpForm()
        return render(request, 'upload_form.html', {"form": form, "title": title})

    form = UpForm(data=request.POST, files=request.FILES)
    if form.is_valid():
        # {'name': '武沛齐', 'age': 123, 'img': <InMemoryUploadedFile: 图片 1.png (image/png)>}
        # 1.读取图片内容，写入到文件夹中并获取文件的路径。
        image_object = form.cleaned_data.get("img")
        # 图片存在项目里面，数据库里存的是地址
        # file_path = "app01/static/img/{}".format(image_object.name)

        # settings.MEDIA_ROOT
        media_path = os.path.join("media", image_object.name)  # 写入文件路径，用来存数据库

        # file_path = os.path.join("app01", db_file_path) # 图片的存储地址
        f = open(media_path, mode='wb')
        for chunk in image_object.chunks():  # 一点点写入
            f.write(chunk)
        f.close()

        # 2.将图片文件路径写入到数据库
        models.Boss.objects.create(
            name=form.cleaned_data['name'],
            age=form.cleaned_data['age'],
            img=media_path,
        )
        return HttpResponse("...")
    return render(request, 'upload_form.html', {"form": form, "title": title})


from app01.utils.bootstrap import BootStrapModelForm


class UpModelForm(BootStrapModelForm):
    bootstrap_exclude_fields = ['img']

    class Meta:
        model = models.City
        fields = "__all__"


def upload_modelform(request):
    """ 上传文件和数据（modelForm）"""
    title = "ModelForm上传文件"
    if request.method == "GET":
        form = UpModelForm()
        return render(request, 'upload_form.html', {"form": form, 'title': title})

    form = UpModelForm(data=request.POST, files=request.FILES)
    if form.is_valid():
        # 对于文件：自动保存；
        # 字段 + 上传路径写入到数据库
        form.save()

        return redirect("/city/list/")
    return render(request, 'upload_form.html', {"form": form, 'title': title})


def city_list(request):
    queryset = models.City.objects.all()
    return render(request, 'city_list.html', {"queryset": queryset})
